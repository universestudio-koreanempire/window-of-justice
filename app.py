import os
from flask import Flask, send_file, jsonify, request
import openpyxl
from collections import Counter
from datetime import datetime
import requests # 새로 추가됨: Gemini API 통신을 위한 HTTP 요청 모듈

app = Flask(__name__, static_folder='static')

@app.route('/')
def home():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    index_path = os.path.join(base_dir, 'index.html')
    return send_file(index_path)

@app.route('/api/trending')
def get_trending():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, 'static', 'searchlist.xlsx')
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        search_terms = [str(row[0]).strip() for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]]
        counter = Counter(search_terms)
        top_3 = counter.most_common(3)
        results = [{'id': idx + 1, 'term': item[0]} for idx, item in enumerate(top_3)]
        return jsonify(results)
    except Exception as e:
        print(f"인기 검색어 엑셀 읽기 오류: {e}")
        return jsonify([])

# [화면 UI 용] 최근 리뷰 3개만 불러오기
@app.route('/api/reviews/recent')
def get_recent_reviews():
    target = request.args.get('target', '')
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, 'static', 'review.xlsx')
    try:
        if not os.path.exists(file_path):
            return jsonify([])
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        reviews = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row): 
                row_target = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                if row_target == str(target).strip():
                    reviews.append({
                        'userId': str(row[0]) if row[0] else '익명',
                        'date': str(row[1]) if row[1] else '',
                        'content': str(row[2]) if row[2] else '',
                        'rating': int(row[3]) if row[3] else 5
                    })
                    if len(reviews) >= 3: # 딱 3개까지만 제한
                        break
        return jsonify(reviews)
    except Exception as e:
        print(f"최근 리뷰 불러오기 오류: {e}")
        return jsonify([])

# [Gemini AI 용] 해당 법관의 '전체 리뷰' 모조리 불러오기
@app.route('/api/reviews/all')
def get_all_reviews():
    target = request.args.get('target', '')
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, 'static', 'review.xlsx')
    try:
        if not os.path.exists(file_path):
            return jsonify([])
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        reviews = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row): 
                row_target = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                if row_target == str(target).strip():
                    reviews.append({
                        'rating': int(row[3]) if row[3] else 5,
                        'content': str(row[2]) if row[2] else ''
                    })
                    # 3개 제한 없이 일치하는 것은 무한정 모두 배열에 담습니다.
        return jsonify(reviews)
    except Exception as e:
        print(f"전체 리뷰 불러오기 오류: {e}")
        return jsonify([])

# [Gemini AI 용] 프론트엔드의 요청을 받아 백엔드에서 리뷰를 모으고 API를 호출하는 로직 (새로 추가)
@app.route('/api/ai/analyze', methods=['POST'])
def analyze_target():
    data = request.json
    target = data.get('targetId', '')
    if not target:
        return jsonify({'success': False, 'message': '대상자가 지정되지 않았습니다.'}), 400

    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, 'static', 'review.xlsx')
    reviews = []
    
    # 1. 엑셀에서 해당 타겟의 전체 리뷰 가져오기 (백엔드 내부 로직)
    try:
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_target = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    if row_target == str(target).strip():
                        reviews.append({
                            'rating': int(row[3]) if row[3] else 5,
                            'content': str(row[2]) if row[2] else ''
                        })
    except Exception as e:
        print(f"리뷰 로드 오류: {e}")

    # 2. 리뷰 컨텍스트 생성
    review_context = ""
    if reviews:
        review_context = "다음은 이 법관에 대해 실제 시민들이 남긴 사이트 내 전체 리뷰 데이터입니다:\n"
        for idx, r in enumerate(reviews):
            review_context += f"{idx + 1}. 별점 {r['rating']}점: \"{r['content']}\"\n"
        review_context += "\n"
    else:
        review_context = "현재 이 법관에 대해 플랫폼에 등록된 시민 리뷰는 없습니다.\n\n"

    # 3. 프롬프트 구성
    prompt_text = f"당신은 '사법의 창'이라는 플랫폼의 수석 AI 법률 데이터 분석관입니다. 사용자가 '{target}'(판사/검사)에 대한 종합 분석을 요청했습니다.\n\n{review_context}위의 데이터(별점 및 내용)를 반드시 분석에 적극적으로 반영하여, 이 법관의 재판 진행 스타일, 특징, 시민들의 평가를 종합한 분석 리포트를 3문장으로 전문성 있게 요약해주세요. (마지막에는 실제 인물이 아닌 테스트 데모임을 명시해주세요.)"

    # 4. Gemini API 호출 (보안을 위해 소스코드 하드코딩 대신 환경 변수에서 API 키를 가져옵니다)
    GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '') 
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-preview-09-2025:generateContent?key={GEMINI_API_KEY}"

    try:
        response = requests.post(
            endpoint,
            headers={'Content-Type': 'application/json'},
            json={"contents": [{"parts": [{"text": prompt_text}]}]}
        )
        response.raise_for_status()
        ai_data = response.json()
        result_text = ai_data.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '분석을 생성하지 못했습니다.')
        return jsonify({'success': True, 'text': result_text})
    except Exception as e:
        print(f"Gemini API 오류: {e}")
        return jsonify({'success': False, 'message': 'AI 분석 서버 통신 중 오류가 발생했습니다.'}), 500


@app.route('/api/review', methods=['POST'])
def save_review():
    try:
        data = request.json
        user_id = data.get('userId', '익명')
        content = data.get('content', '')
        rating = data.get('rating', '5')
        target_id = data.get('targetId', '') 
        
        date_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(base_dir, 'static', 'review.xlsx')
        
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(['아이디', '작성일', '내용', '별점', '대상자(Ax)'])
            
        sheet.insert_rows(2)
        sheet.cell(row=2, column=1, value=user_id)
        sheet.cell(row=2, column=2, value=date_str)
        sheet.cell(row=2, column=3, value=content)
        sheet.cell(row=2, column=4, value=rating)
        sheet.cell(row=2, column=5, value=target_id)
        
        wb.save(file_path)
        
        return jsonify({'success': True, 'message': '리뷰가 저장되었습니다.'})
        
    except Exception as e:
        print(f"리뷰 엑셀 저장 중 오류 발생: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
